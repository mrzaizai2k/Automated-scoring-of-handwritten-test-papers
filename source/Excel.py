import pandas as pd
import os
import time
import cv2
import numpy as np
import matplotlib.pyplot as plt
from create_metrics_OCR import _levenshtein_distance

import openpyxl
from openpyxl import load_workbook

def class_list(dataPath):

    data = pd.read_excel(dataPath)

    MSSV_list = data.iloc[: ,0].to_list()
    Ho_list = data.iloc[: ,1].to_list()
    Ten_list = data.iloc[: ,2].to_list()
    Diem_list = data.iloc[: ,4].to_list()
    name_MSSV_list = []
    name_list = []
    for i in range (len(Ho_list)):
        name_list.append(str(Ho_list[i]+' '+Ten_list[i])) 
        name_MSSV_list.append(str(Ho_list[i]+' '+Ten_list[i] + ' ' + str(MSSV_list[i])))
    return name_list, MSSV_list, name_MSSV_list, Diem_list

def lexicon_search(hypo, dic):
    '''lexicon_search(hypo, dic, method)'''
    dis_list =[]
    index = 0
    for i in range (len(dic)):
        dic[i] = str(dic[i])
        distance  = _levenshtein_distance(hypo.lower(),dic[i].lower())

        dis_list.append(distance)
        if distance == min (dis_list):
            index = i 
    return index, dic[index], distance

def writing_to_excel(path, index, score):
    '''Writing score into Excel file (path, index, score)'''
    wb = openpyxl.load_workbook(path, read_only=False, keep_vba=False)
    ws = wb[wb.sheetnames[0]]
    index = 'F'+str(index)
    ws[index] = score
    wb.save(path)


def main():
    writing_to_excel('data\Class_list.xlsx',4, 3.125)

if __name__ == '__main__':
    main()


